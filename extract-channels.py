#!/usr/bin/env python3
"""
Extract all categories and channels from Xtream API and save to Excel.
Run this on your home network (where the TV is) — the API blocks other IPs.

Usage:  python3 extract-channels.py
Output: channels-export.xlsx
"""

import json, urllib.request, sys
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HOSTS = [
    'my8k.site', 'm3u.best-smarter.me', 'cf.hi-max.me', 'cf.ok4k.me',
    'cf.scma.me', 'hi-ott.me', 'pro.business-cdn.me', 'very50610.cdn-mx.me'
]
USER = '9b1f6b5188'
PASS = '36690c9df5'

def api_call(host, action, extra=''):
    url = f'http://{host}/player_api.php?username={USER}&password={PASS}&action={action}{extra}'
    req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
    with urllib.request.urlopen(req, timeout=15) as resp:
        return json.loads(resp.read())

def find_working_host():
    for host in HOSTS:
        try:
            print(f"  Trying {host}...", end=' ')
            api_call(host, 'get_live_categories')
            print("OK")
            return host
        except Exception as e:
            print(f"failed ({e})")
    return None

def main():
    print("Finding working host...")
    host = find_working_host()
    if not host:
        print("All hosts failed. Make sure you're on your home network.")
        sys.exit(1)

    print(f"\nUsing: {host}")

    print("Fetching categories...")
    categories = api_call(host, 'get_live_categories')
    cat_map = {c['category_id']: c['category_name'] for c in categories}
    print(f"  {len(categories)} categories")

    print("Fetching all channels (this may take a moment)...")
    all_channels = api_call(host, 'get_live_streams')
    print(f"  {len(all_channels)} channels")

    # Build workbook
    wb = Workbook()

    # ── Sheet 1: Categories ──
    ws_cat = wb.active
    ws_cat.title = 'Categories'

    header_font = Font(bold=True, color='FFFFFF', size=11, name='Arial')
    header_fill = PatternFill('solid', fgColor='2B5C3F')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        bottom=Side(style='thin', color='D0D0D0')
    )

    cat_headers = ['Category ID', 'Category Name', 'Channel Count', 'Currently Loaded (ruy)', 'Notes']
    for col, h in enumerate(cat_headers, 1):
        cell = ws_cat.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Count channels per category
    cat_counts = {}
    for ch in all_channels:
        cid = str(ch.get('category_id', ''))
        cat_counts[cid] = cat_counts.get(cid, 0) + 1

    # ruy's current filters (from app.js)
    ruy_includes = ['sport', 'football', 'fotboll', 'soccer', 'futbol',
                    'premier league', 'champions league', 'la liga', 'serie a',
                    'bein', 'dazn', 'espn']

    for i, cat in enumerate(sorted(categories, key=lambda c: c['category_name'].lower()), 2):
        cid = str(cat['category_id'])
        name = cat['category_name']
        count = cat_counts.get(cid, 0)

        # Check if ruy's filters would include this
        name_lower = name.lower()
        included = any(f in name_lower for f in ruy_includes)

        ws_cat.cell(row=i, column=1, value=cid)
        ws_cat.cell(row=i, column=2, value=name)
        ws_cat.cell(row=i, column=3, value=count)
        loaded_cell = ws_cat.cell(row=i, column=4, value='YES' if included else 'no')
        loaded_cell.font = Font(color='00AA00' if included else '999999', name='Arial')
        ws_cat.cell(row=i, column=5, value='')  # Notes column for you

        for col in range(1, 6):
            ws_cat.cell(row=i, column=col).border = thin_border

    ws_cat.column_dimensions['A'].width = 14
    ws_cat.column_dimensions['B'].width = 45
    ws_cat.column_dimensions['C'].width = 16
    ws_cat.column_dimensions['D'].width = 22
    ws_cat.column_dimensions['E'].width = 30
    ws_cat.auto_filter.ref = f'A1:E{len(categories)+1}'
    ws_cat.freeze_panes = 'A2'

    # ── Sheet 2: Channels ──
    ws_ch = wb.create_sheet('Channels')

    ch_headers = ['Channel Name', 'Category', 'Category ID', 'Stream ID', 'EPG Channel ID', 'Added']
    for col, h in enumerate(ch_headers, 1):
        cell = ws_ch.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for i, ch in enumerate(sorted(all_channels, key=lambda c: c.get('name', '').lower()), 2):
        cid = str(ch.get('category_id', ''))
        cat_name = cat_map.get(cid, cat_map.get(ch.get('category_id'), 'Unknown'))
        added = ch.get('added', '')
        if added and added.isdigit():
            try:
                added = datetime.fromtimestamp(int(added)).strftime('%Y-%m-%d')
            except:
                pass

        ws_ch.cell(row=i, column=1, value=ch.get('name', ''))
        ws_ch.cell(row=i, column=2, value=cat_name)
        ws_ch.cell(row=i, column=3, value=cid)
        ws_ch.cell(row=i, column=4, value=str(ch.get('stream_id', '')))
        ws_ch.cell(row=i, column=5, value=ch.get('epg_channel_id', ''))
        ws_ch.cell(row=i, column=6, value=added)

        for col in range(1, 7):
            ws_ch.cell(row=i, column=col).border = thin_border

    ws_ch.column_dimensions['A'].width = 50
    ws_ch.column_dimensions['B'].width = 40
    ws_ch.column_dimensions['C'].width = 14
    ws_ch.column_dimensions['D'].width = 12
    ws_ch.column_dimensions['E'].width = 25
    ws_ch.column_dimensions['F'].width = 14
    ws_ch.auto_filter.ref = f'A1:F{len(all_channels)+1}'
    ws_ch.freeze_panes = 'A2'

    # ── Sheet 3: Summary ──
    ws_sum = wb.create_sheet('Summary')
    ws_sum.sheet_properties.tabColor = '2B5C3F'

    summary_data = [
        ('Export Date', datetime.now().strftime('%Y-%m-%d %H:%M')),
        ('Host', host),
        ('Total Categories', len(categories)),
        ('Total Channels', len(all_channels)),
        ('', ''),
        ('ruy Include Filters', ', '.join(ruy_includes)),
        ('Categories Matching ruy', sum(1 for c in categories if any(f in c['category_name'].lower() for f in ruy_includes))),
    ]
    for i, (label, value) in enumerate(summary_data, 1):
        ws_sum.cell(row=i, column=1, value=label).font = Font(bold=True, name='Arial')
        ws_sum.cell(row=i, column=2, value=value).font = Font(name='Arial')
    ws_sum.column_dimensions['A'].width = 25
    ws_sum.column_dimensions['B'].width = 60

    out = 'channels-export.xlsx'
    wb.save(out)
    print(f"\nSaved to {out}")
    print(f"  - Categories sheet: {len(categories)} rows")
    print(f"  - Channels sheet: {len(all_channels)} rows")
    print(f"  - Summary sheet with filter info")
    print(f"\nOpen it and filter the 'Currently Loaded' column to see what ruy's profile includes.")

if __name__ == '__main__':
    main()
